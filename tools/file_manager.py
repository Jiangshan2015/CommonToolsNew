import openpyxl
from openpyxl_image_loader import SheetImageLoader
from tools.utils import *


def HandleReadFiles(pg, path):
    save_dir = MakeDir('CommonToolsImages')

    file_paths = []
    for root, dirs, files in os.walk(path):
        file_paths = [os.path.join(root, file) for file in files]
    for file_path in file_paths:
        if file_path.endswith('.xlsx'):
            print(file_path)
            workbook = openpyxl.load_workbook(file_path)
            for sheetName in workbook.sheetnames:
                ws = workbook[sheetName]
                image_loader = SheetImageLoader(ws)
                num = ws.max_row
                for i in range(2, num + 1):  # 从第2行开始，总行数要+1
                    try:
                        name = ws['B' + str(i)].value  # B列的文件名
                        image = image_loader.get('C' + str(i))  # C列的图片
                        image = image.convert('RGB')
                        image_name = save_dir + '/' + name + ".jpg"
                        image.save(image_name)  # 以Ai为名，存图片Ci
                    # 排除没有图片，或图片超出单元格的情况
                    except ValueError:
                        print("这一行没有图片：", i)
