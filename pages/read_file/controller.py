import os
import openpyxl
from flet_core import FilePickerResultEvent
from flet_mvc import FletController, alert
from openpyxl_image_loader import SheetImageLoader
import sys  # 添加 sys 模块导入
from pages.read_file.widget import top_tip_widget, loading_widget, start_read_widget
from tools.utils import MakeDir
from views.choose_file_view import ChooseFileView
from time import sleep


# Controller
class ReadFileController(FletController):
    # 选择文件按钮
    def file_on_click(self, e=None):
        self.model.get_directory_dialog.value.get_directory_path()

    # 选完文件回调
    def get_directory_result(self, e: FilePickerResultEvent):
        self.model.dir_path.value = e.path if e.path else ""
        if self.model.dir_path.value == "":
            self.alert("你未选中任何文件夹", alert.WARNING)
        else:
            dir_name = os.path.basename(self.model.dir_path.value)
            self.model.read_file_content.set_value([top_tip_widget(),
                                                    ChooseFileView(on_click=self.file_on_click, pathName=dir_name),
                                                    start_read_widget(click=self.start_read),
                                                    ])
        self.update()

    # 开始读取
    def start_read(self, e=None):
        if self.model.dir_path.value == "":
            self.alert("你未选中任何文件夹", alert.WARNING)
        else:
            dir_name = os.path.basename(self.model.dir_path.value)
            self.model.read_file_content.set_value([top_tip_widget(),
                                                    ChooseFileView(on_click=self.file_on_click, pathName=dir_name),
                                                    loading_widget(),
                                                    ])
        self.update()
        self.handleFiles()
        sleep(3)
        self.model.dir_path.value = ""
        self.model.read_file_content.set_value([top_tip_widget(),
                                                ChooseFileView(on_click=self.file_on_click),
                                                start_read_widget(click=self.start_read),
                                                ])
        self.update()

    def handleFiles(self):
        if self.model.dir_path.value == "":
            return
        # 使用 os.path 处理桌面路径
        if sys.platform == 'win32':  # Windows
            img_dir = os.path.join(os.path.expanduser('~'), 'Desktop', 'CommonToolsImages')
        else:  # macOS 或 Linux
            img_dir = os.path.expanduser('~/Desktop/CommonToolsImages')
            
        if not os.path.exists(img_dir):
            os.makedirs(img_dir)
            
        file_paths = []
        for root, dirs, files in os.walk(self.model.dir_path.value):
            # 过滤隐藏文件和非xlsx文件
            valid_files = [file for file in files 
                         if not file.startswith('.') and  # 排除 Unix/Mac 隐藏文件
                         not file.startswith('~$') and    # 排除 Excel 临时文件
                         file.lower().endswith('.xlsx')]  # 只包含 xlsx 文件
            file_paths.extend(os.path.join(root, file) for file in valid_files)
            
        for file_path in file_paths:
            workbook = openpyxl.load_workbook(file_path)
            for sheetName in workbook.sheetnames:
                ws = workbook[sheetName]
                image_loader = SheetImageLoader(ws)
                num = ws.max_row
                for i in range(2, num + 1):  # 从第2行开始，总行数要+1
                    try:
                        name = ws['A' + str(i)].value  # B列的文件名
                        if name:
                            # 替换文件名中的路径分隔符
                            safe_name = name.replace('/', '*_*').replace('\\', '_')
                            image = image_loader.get('B' + str(i))  # C列的图片
                            image = image.convert('RGB')
                            # 使用安全的文件名构建保存路径
                            image_name = os.path.join(img_dir, f"{safe_name}.jpg")
                            image.save(image_name)
                    # 排除没有图片，或图片超出单元格的情况
                    except ValueError:
                        print("这一行没有图片：", i)

        print("结束了")
