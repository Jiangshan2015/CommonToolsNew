import os
import openpyxl
from flet_core import FilePickerResultEvent
from flet_mvc import FletController, alert
from openpyxl.styles import PatternFill

from pages.write_file.widget import top_tip_widget, loading_widget, start_write_widget
from views.choose_file_view import ChooseFileView
from openpyxl.drawing.image import Image
from time import sleep
import sys


# Controller
class WriteFileController(FletController):
    # 选择文件按钮
    def file_on_click(self, e=None):
        self.model.get_directory_dialog.value.get_directory_path()

    # 选完文件回调
    def get_directory_result(self, e: FilePickerResultEvent):
        self.model.dir_path.value = e.path if e.path else ""
        if self.model.dir_path.value == "":
            self.alert("你未选中任何文件夹", alert.WARNING)
        else:
            # arr = self.model.dir_path.value.split('/')
            path_name = os.path.basename(self.model.dir_path.value)
            self.model.write_file_content.set_value([top_tip_widget(),
                                                    ChooseFileView(on_click=self.file_on_click, pathName=path_name),
                                                    start_write_widget(click=self.start_write),
                                                    ])
        self.update()

    # 开始读取
    def start_write(self, e=None):
        if self.model.dir_path.value == "":
            self.alert("你未选中任何文件夹", alert.WARNING)
        else:
            path_name = os.path.basename(self.model.dir_path.value)
            self.model.write_file_content.set_value([top_tip_widget(),
                                                    ChooseFileView(on_click=self.file_on_click, pathName=path_name),
                                                    loading_widget(),
                                                    ])
        self.update()
        self.handleFiles()
        sleep(3)
        self.model.dir_path.value = ""
        self.model.write_file_content.set_value([top_tip_widget(),
                                                ChooseFileView(on_click=self.file_on_click),
                                                start_write_widget(click=self.start_write),
                                                ])
        self.update()

    def handleFiles(self):
        file_paths = []
        img_name_column = 'B'
        img_column = 'C'
        # 使用 os.path 处理桌面路径
        if sys.platform == 'win32':  # Windows
            img_dir = os.path.join(os.path.expanduser('~'), 'Desktop', 'CommonToolsImages')
        else:  # macOS 或 Linux
            img_dir = os.path.expanduser('~/Desktop/CommonToolsImages')
        
        # 修改文件遍历逻辑，过滤隐藏文件和非xlsx文件
        for root, dirs, files in os.walk(self.model.dir_path.value):
            valid_files = [file for file in files 
                         if file.endswith('.xlsx') 
                         and not file.startswith('.') 
                         and not file.startswith('~$')]
            file_paths.extend([os.path.join(root, file) for file in valid_files])
        
        for file_path in file_paths:
            try:
                print(f"正在处理文件: {file_path}")
                workbook = openpyxl.load_workbook(file_path)
                # 获取sheet页
                for sheetName in workbook.sheetnames:
                    try:
                        ws = workbook[sheetName]
                        print(f"正在处理工作表: {sheetName}")
                        num = ws.max_row
                        for i in range(1, num):  # 从第2行开始，总行数要+1
                            img_file_path = ''
                            img_name = ws[img_name_column + str(i)].value
                            if img_name is not None:
                                img_file_path = img_dir + '/' + img_name + '.jpg'
                            if os.path.exists(img_file_path):
                                try:
                                    # 检查单元格是否已经有图片
                                    cell_has_image = False
                                    for image in ws._images:
                                        if image.anchor == f"{img_column}{i}":
                                            cell_has_image = True
                                            break
                                    if not cell_has_image:
                                        # 获取图片
                                        img = Image(img_file_path)
                                        # 设置图片的大小为100x100像素
                                        img.width, img.height = (100, 100)
                                        # 设置表格的宽度和高度以匹配图片尺寸
                                        # 列宽：100像素 ≈ 13个单位（100/8 ≈ 13）
                                        ws.column_dimensions[img_column].width = 13
                                        # 行高：100像素 ≈ 133个单位（100/0.75 ≈ 133）
                                        ws.row_dimensions[i].height = 133
                                        # 图片插入名称对应单元格
                                        ws.add_image(img, f"{img_column}{i}")
                                    else:
                                        print(f"单元格 {img_column}{i} 已存在图片，跳过")
                                except Exception as e:
                                    print(f"插入图片失败: {img_file_path}")
                                    print(f"错误信息: {str(e)}")
                                    ws.cell(row=i, column=1).fill = PatternFill(patternType='solid', fgColor='FF0000')  # 红色标记错误
                            else:
                                print(f"未找到对应图片---->{img_file_path}")
                                ws.cell(row=i,column=1).fill = PatternFill(patternType='solid',fgColor='90EE90')#淡绿色
                    except Exception as e:
                        print(f"处理工作表 {sheetName} 时出错: {str(e)}")
                        continue
                workbook.save(file_path)
                workbook.close()
                print(f'文件 {file_path} 处理完成')
            except Exception as e:
                print(f"无法打开或处理文件 {file_path}: {str(e)}")
                continue

        print("所有文件处理完成")
